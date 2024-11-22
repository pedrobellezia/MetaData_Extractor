"""Esse modulo processa arquivos de mídia para extrair metadados, criar relatórios Excel e gerar gráficos."""
import json
import re
import logging
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from pymediainfo import MediaInfo
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Configuração do logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def planificar_diretorio(path):
    """Planifica a estrutura de diretórios."""
    path = Path(path)
    stack = [path]
    itens = {}

    # Itera sobre os itens no diretório
    while stack:
        current = stack.pop()
        for item in current.iterdir():
            if item.is_dir():
                stack.append(item)
            elif item.is_file():
                itens[str(item)] = item.name

    return itens

def format_size(size):
    """Formata o tamanho do arquivo."""
    size = int(size)
    return f"{size / 1_000_000:.2f} MB"

def teste(arqui):
    """Extrai os metadados do arquivo."""
    arq = arqui[0]

    replace = r'\\192\.168\.4\.71\\'
    path = re.sub(replace, "", arqui[0])
    try:
        media_info = MediaInfo.parse(arq)
    except Exception as e:
        logging.error(f"Erro ao analisar o arquivo {arq}: {e}")
        return None

    resultado = {
        "nome_arquivo": arqui[1],
        "data_criacao": None,
        "data_modi": None,
        "duracao": None,
        "size": None,
        "caminho": f"{path}",
        "extensao": Path(arqui[1]).suffix
    }

    date_pattern = r'\d{4}-\d{2}-\d{2}'

    # Extrai informações dos metadados
    for track in media_info.tracks:
        if track.track_type == "General":
            if track.file_size:
                resultado["size"] = format_size(track.file_size)

        if track.encoded_date:
            match = re.search(date_pattern, track.encoded_date)
            if match:
                resultado["data_criacao"] = match.group(0)

        if track.tagged_date:
            match = re.search(date_pattern, track.tagged_date)
            if match:
                resultado["data_modi"] = match.group(0)

        if track.duration:
            duracao_em_ms = int(track.duration)

            horas = duracao_em_ms // (1000 * 60 * 60)
            minutos = (duracao_em_ms // (1000 * 60)) % 60
            segundos = (duracao_em_ms // 1000) % 60

            resultado["duracao"] = f"{horas:02}:{minutos:02}:{segundos:02}"
    logging.info(resultado)
    return resultado

def processar_diretorio(direc):
    """Processa o diretório para extrair metadados de todos os arquivos."""
    final_j = []

    # Itera sobre os diretórios fornecidos
    for d in direc:
        try:
            resultado_json = list(planificar_diretorio(d).items())
        except Exception as e:
            logging.error(f"Erro ao planificar o diretório {d}: {e}")
            continue

        for arquivo in resultado_json:
            new_dic = teste(arquivo)
            if new_dic:
                final_j.append(new_dic)

    return final_j

def create_excel(dados, excel_name):
    """Cria ou atualiza um arquivo Excel com os dados fornecidos."""
    df = pd.DataFrame.from_dict(dados)
    excel_file = excel_name

    try:
        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            for sheet_name in writer.book.sheetnames:
                start_row = writer.book[sheet_name].max_row
                df.to_excel(writer, index=False, header=False, startrow=start_row, sheet_name=sheet_name)
                return
    except FileNotFoundError:
        df.to_excel(excel_file, index=False)
    except Exception as e:
        logging.error(f"Erro ao criar ou atualizar o arquivo Excel {excel_file}: {e}")

def create_extension_graph(xl):
    """Cria um gráfico de barras das extensões de arquivo."""
    try:
        df = pd.read_excel(xl)
    except Exception as e:
        logging.error(f"Erro ao ler o arquivo Excel {xl}: {e}")
        return

    df['extensao'] = df['extensao'].str.lower()
    contagem_extensoes = df['extensao'].value_counts()
    plt.figure(figsize=(10, 6))
    contagem_extensoes.plot(kind='bar')  # kind = tipo de gráfico
    plt.xlabel('Extensão')  # nome do eixo x
    plt.ylabel('Contagem')  # nome do eixo y
    plt.title('Distribuição das Extensões')  # título do gráfico
    plt.xticks(rotation=45)  # rotaciona os rótulos do eixo x
    plt.savefig('graph.png')
    plt.close()

def insert_image(xl, imge, column):
    """Insere uma imagem no arquivo Excel na coluna especificada."""
    try:
        wb = load_workbook(xl)
        ws = wb.active
        img = Image(imge)
        ws.add_image(img, column)
        wb.save(xl)
    except Exception as e:
        logging.error(f"Erro ao inserir a imagem {imge} no arquivo Excel {xl}: {e}")

def save_to_json(dados, json_name):
    """Salva os dados fornecidos em um arquivo JSON."""
    try:
        with open(json_name, 'w', encoding='utf-8') as json_file:
            json.dump(dados, json_file, ensure_ascii=False, indent=4)
    except Exception as e:
        logging.error(f"Erro ao salvar os dados no arquivo JSON {json_name}: {e}")



# Diretórios a serem processados
DIRETORIO = [r"C:\Users\pedrobs\Documents\tvcmsj"]
# Nome do arquivo Excel
EXCEL_NAME = "xlsxteste.xlsx"
JSON_NAME = "jsonteste.json"
# Processa os diretórios e cria o relatório Excel
final_result = processar_diretorio(DIRETORIO)
create_excel(final_result, EXCEL_NAME)
save_to_json(final_result, JSON_NAME)

# Cria o gráfico de extensões e insere no Excel
create_extension_graph(EXCEL_NAME)
insert_image(EXCEL_NAME, "graph.png", "H1")